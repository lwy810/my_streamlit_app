import requests
import pandas as pd
import json
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import traceback
import urllib.parse

class AirQualityReporter:
    def __init__(self, api_key):
        """
        미세먼지 리포터 초기화
        
        Args:
            api_key (str): 에어코리아 Open API 키
        """
        self.api_key = api_key
        self.base_url = "http://apis.data.go.kr/B552584/ArpltnInforInqireSvc"
        
    def get_sido_list(self):
        """시도 목록 조회"""
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': self.api_key,
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': '서울',
            'ver': '1.0'
        }
        
        try:
            response = requests.get(url, params=params)
            if response.status_code == 200:
                # 주요 시도 목록 (실제 API에서는 동적으로 가져올 수 있음)
                sido_list = [
                    '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                    '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주'
                ]
                return sido_list
        except Exception as e:
            print(f"시도 목록 조회 오류: {e}")
            return []
    
    def get_realtime_air_quality(self, sido_name):
        """실시간 대기질 데이터 조회"""
        # 먼저 XML로 시도
        xml_data = self._fetch_air_quality_xml(sido_name)
        if xml_data:
            return xml_data
        
        # XML 실패 시 JSON으로 시도
        return self._fetch_air_quality_json(sido_name)
    
    def _fetch_air_quality_xml(self, sido_name):
        """XML 형태로 대기질 데이터 조회"""
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': self.api_key,
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': sido_name,
            'ver': '1.0'
        }
        
        try:
            print(f"[XML] API 요청 URL: {url}")
            print(f"[XML] API 요청 파라미터: {params}")
            
            response = requests.get(url, params=params, timeout=30)
            print(f"[XML] 응답 상태 코드: {response.status_code}")
            print(f"[XML] Content-Type: {response.headers.get('content-type', 'Unknown')}")
            print(f"[XML] 응답 내용 (처음 300자): {response.text[:300]}")
            
            if response.status_code == 200:
                return self._parse_xml_response(response.text, sido_name)
            else:
                print(f"[XML] API 요청 실패. 상태 코드: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"[XML] 요청 오류: {e}")
            return None
    
    def _fetch_air_quality_json(self, sido_name):
        """JSON 형태로 대기질 데이터 조회"""
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': self.api_key,
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': sido_name,
            'ver': '1.0'
        }
        
        try:
            print(f"[JSON] API 요청 URL: {url}")
            print(f"[JSON] API 요청 파라미터: {params}")
            
            response = requests.get(url, params=params, timeout=30)
            print(f"[JSON] 응답 상태 코드: {response.status_code}")
            print(f"[JSON] Content-Type: {response.headers.get('content-type', 'Unknown')}")
            print(f"[JSON] 응답 내용 (처음 300자): {response.text[:300]}")
            
            if response.status_code != 200:
                print(f"[JSON] API 요청 실패. 상태 코드: {response.status_code}")
                return None
            
            # 빈 응답 체크
            if not response.text.strip():
                print("[JSON] 빈 응답 수신")
                return None
            
            # JSON 파싱 시도
            try:
                data = response.json()
                print(f"[JSON] 파싱 성공: {type(data)}")
                
                if isinstance(data, dict) and 'response' in data:
                    response_data = data['response']
                    if 'body' in response_data and 'items' in response_data['body']:
                        items = response_data['body']['items']
                        if items:
                            print(f"[JSON] 아이템 수: {len(items)}")
                            return self._process_air_quality_data(items)
                        else:
                            print("[JSON] 아이템이 비어있음")
                    else:
                        print(f"[JSON] body/items가 없음: {response_data.keys()}")
                else:
                    print(f"[JSON] 예상과 다른 응답 구조: {type(data)}")
                
                return None
                
            except json.JSONDecodeError as json_err:
                print(f"[JSON] JSON 파싱 오류: {json_err}")
                print(f"[JSON] 응답 전체 내용: {response.text}")
                return None
                
        except requests.exceptions.Timeout:
            print("[JSON] 요청 타임아웃")
            return None
        except requests.exceptions.RequestException as req_err:
            print(f"[JSON] 네트워크 요청 오류: {req_err}")
            return None
        except Exception as e:
            print(f"[JSON] 알 수 없는 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _parse_xml_response(self, xml_text, sido_name):
        """XML 응답 파싱"""
        try:
            print(f"[XML Parser] XML 파싱 시작")
            print(f"[XML Parser] XML 내용 (처음 500자): {xml_text[:500]}")
            
            root = ET.fromstring(xml_text)
            print(f"[XML Parser] 루트 요소: {root.tag}")
            
            # 네임스페이스 처리
            namespaces = dict([node for _, node in ET.iterparse(BytesIO(xml_text.encode()), events=['start-ns'])])
            print(f"[XML Parser] 네임스페이스: {namespaces}")
            
            # 에러 코드 확인 (여러 가능한 경로)
            result_code_paths = ['.//resultCode', './/header/resultCode', './/{*}resultCode']
            result_code = None
            for path in result_code_paths:
                result_code = root.find(path)
                if result_code is not None:
                    break
            
            if result_code is not None:
                code = result_code.text
                print(f"[XML Parser] 결과 코드: {code}")
                
                if code != '00':
                    result_msg_paths = ['.//resultMsg', './/header/resultMsg', './/{*}resultMsg']
                    result_msg = None
                    for path in result_msg_paths:
                        result_msg = root.find(path)
                        if result_msg is not None:
                            break
                    
                    error_msg = result_msg.text if result_msg is not None else "알 수 없는 오류"
                    print(f"[XML Parser] API 에러: {code} - {error_msg}")
                    return None
            
            # items 데이터 추출 (여러 가능한 경로)
            items_paths = ['.//items/item', './/item', './/{*}items/{*}item', './/{*}item']
            items = []
            
            for path in items_paths:
                found_items = root.findall(path)
                if found_items:
                    print(f"[XML Parser] {len(found_items)}개 아이템을 경로 '{path}'에서 발견")
                    for item in found_items:
                        item_data = {}
                        for child in item:
                            # 네임스페이스 제거한 태그명 사용
                            tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                            item_data[tag_name] = child.text
                        items.append(item_data)
                    break
            
            print(f"[XML Parser] 총 {len(items)}개 아이템 추출")
            
            if items:
                # 첫 번째 아이템 샘플 출력
                if len(items) > 0:
                    print(f"[XML Parser] 첫 번째 아이템 키: {list(items[0].keys())}")
                return self._process_air_quality_data(items)
            else:
                print(f"[XML Parser] '{sido_name}' 지역의 데이터를 찾을 수 없습니다.")
                # XML 구조 디버깅
                print("[XML Parser] XML 구조 분석:")
                self._debug_xml_structure(root, 0, 3)  # 최대 3레벨까지 출력
                return None
                
        except ET.ParseError as parse_err:
            print(f"[XML Parser] XML 파싱 오류: {parse_err}")
            print(f"[XML Parser] 오류 위치 근처 내용: {xml_text[max(0, parse_err.position[1]-100):parse_err.position[1]+100]}")
            return None
        except Exception as e:
            print(f"[XML Parser] XML 응답 처리 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _debug_xml_structure(self, element, level, max_level):
        """XML 구조 디버깅용 메소드"""
        if level > max_level:
            return
        
        indent = "  " * level
        tag_name = element.tag.split('}')[-1] if '}' in element.tag else element.tag
        
        if element.text and element.text.strip():
            print(f"{indent}{tag_name}: {element.text.strip()}")
        else:
            print(f"{indent}{tag_name}")
        
        for child in list(element)[:3]:  # 각 레벨에서 최대 3개 자식만 출력
            self._debug_xml_structure(child, level + 1, max_level)

    def test_api_connection(self):
        """API 연결 테스트"""
        print("=" * 50)
        print("API 연결 테스트 시작")
        print("=" * 50)
        
        # 1. 기본 연결 테스트
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        test_params = {
            'serviceKey': self.api_key,
            'numOfRows': '1',
            'pageNo': '1',
            'sidoName': '서울',
            'ver': '1.0'
        }
        
        try:
            print(f"테스트 URL: {url}")
            print(f"테스트 파라미터: {test_params}")
            
            response = requests.get(url, params=test_params, timeout=10)
            print(f"응답 상태 코드: {response.status_code}")
            print(f"응답 헤더: {dict(response.headers)}")
            print(f"응답 내용 (처음 500자):\n{response.text[:500]}")
            
            if response.status_code == 200:
                # API 키 유효성 확인
                if 'INVALID' in response.text.upper() or 'ERROR' in response.text.upper():
                    print("⚠️ API 키가 유효하지 않을 가능성이 있습니다.")
                    return False
                
                # XML 파싱 테스트
                if response.text.strip().startswith('<?xml') or '<' in response.text:
                    try:
                        root = ET.fromstring(response.text)
                        result_code = root.find('.//resultCode')
                        if result_code is not None:
                            if result_code.text == '00':
                                print("✅ API 연결 및 인증 성공!")
                                return True
                            else:
                                result_msg = root.find('.//resultMsg')
                                error_msg = result_msg.text if result_msg is not None else "알 수 없는 오류"
                                print(f"❌ API 오류: {result_code.text} - {error_msg}")
                                return False
                        else:
                            print("⚠️ 결과 코드를 찾을 수 없습니다.")
                            return False
                    except ET.ParseError as e:
                        print(f"❌ XML 파싱 오류: {e}")
                        return False
                else:
                    print("⚠️ 예상과 다른 응답 형식입니다.")
                    return False
            else:
                print(f"❌ HTTP 오류: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            print("❌ 요청 타임아웃")
            return False
        except requests.exceptions.RequestException as e:
            print(f"❌ 네트워크 오류: {e}")
            return False
        except Exception as e:
            print(f"❌ 알 수 없는 오류: {e}")
            import traceback
            traceback.print_exc()
            return False
        """대기질 예보 정보 조회"""
        url = f"{self.base_url}/getMinuDustFrcstDspth"
        params = {
            'serviceKey': self.api_key,
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'searchDate': search_date,
            'InformCode': 'PM10'
        }
        
        try:
            response = requests.get(url, params=params)
            if response.status_code == 200:
                data = response.json()
                if 'response' in data and 'body' in data['response']:
                    items = data['response']['body']['items']
                    return items
            return None
        except Exception as e:
            print(f"대기질 예보 조회 오류: {e}")
            return None
    
    def _process_air_quality_data(self, items):
        """대기질 데이터 처리"""
        processed_data = []
        for item in items:
            processed_item = {
                'stationName': item.get('stationName', ''),
                'sidoName': item.get('sidoName', ''),
                'pm10Value': item.get('pm10Value', ''),
                'pm25Value': item.get('pm25Value', ''),
                'pm10Grade': item.get('pm10Grade', ''),
                'pm25Grade': item.get('pm25Grade', ''),
                'so2Value': item.get('so2Value', ''),
                'coValue': item.get('coValue', ''),
                'o3Value': item.get('o3Value', ''),
                'no2Value': item.get('no2Value', ''),
                'dataTime': item.get('dataTime', ''),
                'khaiValue': item.get('khaiValue', ''),
                'khaiGrade': item.get('khaiGrade', '')
            }
            processed_data.append(processed_item)
        return processed_data
    
    def get_air_quality_grade_text(self, grade):
        """대기질 등급을 텍스트로 변환"""
        grade_map = {
            '1': '좋음',
            '2': '보통',
            '3': '나쁨',
            '4': '매우나쁨'
        }
        return grade_map.get(str(grade), '정보없음')
    
    def create_pdf_report(self, data, sido_name, output_path):
        """PDF 리포트 생성"""
        # 한글 폰트 등록
        font_name = register_korean_font()
        styles = getSampleStyleSheet()
        if font_name:
            from reportlab.lib.styles import ParagraphStyle
            styles.add(ParagraphStyle(name='KoreanTitle', fontName=font_name, fontSize=18, leading=22))
            styles.add(ParagraphStyle(name='KoreanNormal', fontName=font_name, fontSize=10, leading=14))
            title_style = styles['KoreanTitle']
            normal_style = styles['KoreanNormal']
        else:
            title_style = styles['Title']
            normal_style = styles['Normal']

        doc = SimpleDocTemplate(output_path, pagesize=A4)
        story = []
        
        # 제목
        title = Paragraph(f"{sido_name} 지역 미세먼지 현황 보고서", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # 생성 시간
        now = datetime.now().strftime("%Y년 %m월 %d일 %H시 %M분")
        date_para = Paragraph(f"생성일시: {now}", normal_style)
        story.append(date_para)
        story.append(Spacer(1, 20))
        
        # 데이터가 있는 경우에만 테이블 생성
        if data:
            # 테이블 데이터 준비
            table_data = [['측정소명', 'PM10', 'PM2.5', 'PM10등급', 'PM2.5등급', '측정시간']]
            
            for item in data[:10]:  # 상위 10개 측정소만 표시
                table_data.append([
                    item['stationName'],
                    f"{item['pm10Value']}㎍/㎥" if item['pm10Value'] else '-',
                    f"{item['pm25Value']}㎍/㎥" if item['pm25Value'] else '-',
                    self.get_air_quality_grade_text(item['pm10Grade']),
                    self.get_air_quality_grade_text(item['pm25Grade']),
                    item['dataTime']
                ])
            
            # 테이블 생성
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name if font_name else 'Helvetica'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name if font_name else 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
        else:
            no_data_para = Paragraph("데이터를 불러올 수 없습니다.", normal_style)
            story.append(no_data_para)
        
        # PDF 생성
        doc.build(story)
        return output_path
    
    def create_excel_report(self, data, sido_name, output_path):
        """Excel 리포트 생성"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{sido_name} 미세먼지 현황"
        
        # 헤더 설정
        headers = ['측정소명', '시도명', 'PM10 농도', 'PM2.5 농도', 'PM10 등급', 'PM2.5 등급', 
                  'SO2', 'CO', 'O3', 'NO2', '측정시간', '통합대기환경지수', '통합등급']
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        if data:
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item['stationName'])
                ws.cell(row=row, column=2, value=item['sidoName'])
                ws.cell(row=row, column=3, value=item['pm10Value'])
                ws.cell(row=row, column=4, value=item['pm25Value'])
                ws.cell(row=row, column=5, value=self.get_air_quality_grade_text(item['pm10Grade']))
                ws.cell(row=row, column=6, value=self.get_air_quality_grade_text(item['pm25Grade']))
                ws.cell(row=row, column=7, value=item['so2Value'])
                ws.cell(row=row, column=8, value=item['coValue'])
                ws.cell(row=row, column=9, value=item['o3Value'])
                ws.cell(row=row, column=10, value=item['no2Value'])
                ws.cell(row=row, column=11, value=item['dataTime'])
                ws.cell(row=row, column=12, value=item['khaiValue'])
                ws.cell(row=row, column=13, value=self.get_air_quality_grade_text(item['khaiGrade']))
        
        # 열 너비 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        return output_path
    
    def send_email_report(self, sender_email, sender_password, recipient_emails, 
                         subject, body, attachment_paths):
        """이메일로 리포트 발송"""
        try:
            # SMTP 서버 설정 (Gmail 기준)
            smtp_server = "smtp.gmail.com"
            smtp_port = 587
            
            # 이메일 메시지 생성
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = ", ".join(recipient_emails)
            msg['Subject'] = subject
            
            # 본문 추가
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
            
            # 첨부파일 추가
            for file_path in attachment_paths:
                if os.path.exists(file_path):
                    with open(file_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                    
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {os.path.basename(file_path)}'
                    )
                    msg.attach(part)
            
            # 이메일 발송
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(sender_email, sender_password)
            text = msg.as_string()
            server.sendmail(sender_email, recipient_emails, text.encode('utf-8'))
            server.quit()
            
            return True
        except Exception as e:
            print(f"이메일 발송 오류: {e}")
            return False

# Streamlit UI 구현
API_KEY = "SpBjtAWYaN2aNqknzNlYA4wmB0Amo1IcAM8cNfrU5NAk8nuKEtNGw5dNf6MtkVwliAuKWek+4YG8zjq+osj2og=="  # 여기에 본인의 API 키를 입력하세요

FONT_PATHS = [
    "C:/Windows/Fonts/malgun.ttf",
    "./fonts/NanumGothic.ttf",
    "./NanumGothic.ttf"
]

def register_korean_font():
    for font_path in FONT_PATHS:
        if os.path.exists(font_path):
            pdfmetrics.registerFont(TTFont('KoreanFont', font_path))
            return 'KoreanFont'
    return None

def main():
    st.set_page_config(
        page_title="미세먼지 현황 리포트",
        page_icon="🌫️",
        layout="wide"
    )
    
    st.title("🌫️ 전국 미세먼지 농도 현황 및 리포트 생성")
    st.markdown("---")
    
    # st.session_state 초기화
    if 'report_files' not in st.session_state:
        st.session_state.report_files = []
    if 'air_quality_data' not in st.session_state:
        st.session_state.air_quality_data = None
    if 'selected_sido_for_data' not in st.session_state:
        st.session_state.selected_sido_for_data = None
    if 'report_generated_success' not in st.session_state: # **새로 추가된 부분**
        st.session_state.report_generated_success = False   # **리포트 생성 성공 여부를 저장**

    # 사이드바에서 설정
    with st.sidebar:
        st.header("설정")
        
        # reporter는 하드코딩된 API_KEY로 생성
        reporter = AirQualityReporter(API_KEY)
        
        # 'API 연결 테스트' 버튼 및 관련 코드 제거
        
        # 지역 선택
        sido_list = reporter.get_sido_list()
        selected_sido = st.selectbox("지역 선택", sido_list)
        
        # 리포트 타입 선택
        report_types = st.multiselect(
            "생성할 리포트 타입",
            ["PDF", "Excel"],
            default=["PDF"]
        )
        
        # 이메일 설정
        st.subheader("이메일 발송 설정")
        send_email = st.checkbox("이메일로 리포트 발송")
        
        if send_email:
            sender_email = st.text_input("발송자 이메일")
            sender_password = st.text_input("이메일 비밀번호", type="password")
            recipient_emails = st.text_input(
                "수신자 이메일 (콤마로 구분)",
                placeholder="email1@example.com, email2@example.com"
            )
    
    # 메인 콘텐츠
    col1, col2 = st.columns([2, 1])

    with col2:
        if st.button("📊 데이터 조회 및 리포트 생성", type="primary"):
            # <<<<<<< 변경 시작: 리포트 생성 성공 상태 초기화 >>>>>>>
            st.session_state.report_generated_success = False # **새로운 리포트 생성 시도 시 상태 초기화**
            # <<<<<<< 변경 끝: 리포트 생성 성공 상태 초기화 >>>>>>>


            if not selected_sido:
                st.error("지역을 선택해주세요")
                return
            
            # 프로그레스 바
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # 데이터 조회
            status_text.text("실시간 대기질 데이터 조회 중...")
            progress_bar.progress(25)
            
            air_quality_data = reporter.get_realtime_air_quality(selected_sido)
            
            if not air_quality_data:
                st.error("❌ 데이터를 불러올 수 없습니다.")
                st.info("🔍 문제 해결 방법:\n1. API 키가 올바른지 확인\n2. 네트워크 연결 상태 확인\n3. API 키 승인 상태 확인\n4. 다른 지역 선택해보기")
                 # 실패 시 session_state 초기화
                st.session_state.report_files = []
                st.session_state.air_quality_data = None
                # <<<<<<< 변경 시작: 실패 시 성공 상태 초기화 >>>>>>>
                st.session_state.report_generated_success = False # **리포트 생성 실패 시 성공 상태 초기화**
                # <<<<<<< 변경 끝: 실패 시 성공 상태 초기화 >>>>>>>
                return
            
            # 리포트 생성
            status_text.text("리포트 생성 중...")
            progress_bar.progress(50)
            
            new_report_files = []

            # PDF 리포트 생성
            if "PDF" in report_types:
                pdf_path = f"{selected_sido}_미세먼지_리포트_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                reporter.create_pdf_report(air_quality_data, selected_sido, pdf_path)
                new_report_files.append(pdf_path)
            
            # Excel 리포트 생성
            if "Excel" in report_types:
                excel_path = f"{selected_sido}_미세먼지_리포트_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                reporter.create_excel_report(air_quality_data, selected_sido, excel_path)
                new_report_files.append(excel_path)
        

            # session_state에 생성된 파일 목록 업데이트
            st.session_state.report_files = new_report_files
            
            progress_bar.progress(75)

            # 이메일 발송
            if send_email and sender_email and sender_password and recipient_emails:
                status_text.text("이메일 발송 중...")
                recipient_list = [email.strip() for email in recipient_emails.split(',')]
                
                subject = f"[{selected_sido}] 미세먼지 현황 리포트 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                body = f"""
안녕하세요.

{selected_sido} 지역의 미세먼지 현황 리포트를 첨부파일로 보내드립니다.

생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H시 %M분')}
대상지역: {selected_sido}
측정소 수: {len(air_quality_data)}개

감사합니다.
                """
                
                email_result = reporter.send_email_report(
                    sender_email, sender_password, recipient_list,
                    subject, body, st.session_state.report_files
                )
                
                if email_result:
                    st.success("✅ 이메일이 성공적으로 발송되었습니다!")
                else:
                    st.error("❌ 이메일 발송에 실패했습니다.")
            
            progress_bar.progress(100)
            status_text.text("완료!")
            
            # <<<<<<< 변경 시작: 리포트 생성 성공 상태 설정 >>>>>>>
            # 리포트 파일이 실제로 생성되었을 때만 성공 메시지를 표시하고 상태를 True로 설정
            if st.session_state.report_files:
                st.session_state.report_generated_success = True 
            # <<<<<<< 변경 끝: 리포트 생성 성공 상태 설정 >>>>>>>

        # <<<<<<< 변경 시작: 리포트 생성 성공 메시지 지속적으로 표시 >>>>>>>
        # report_generated_success 상태가 True이고, 리포트 파일이 있을 경우 메시지 표시
        if st.session_state.report_generated_success and st.session_state.report_files:
            st.success(f"✅ 리포트가 성공적으로 생성되었습니다!")
        # <<<<<<< 변경 끝: 리포트 생성 성공 메시지 지속적으로 표시 >>>>>>>

        # 다운로드 버튼 (session_state에 파일이 있을 경우 항상 표시)
        if st.session_state.report_files:
            st.subheader("⬇️ 리포트 다운로드")
            for file_path in st.session_state.report_files:
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as f:
                        st.download_button(
                            label=f"📁 {os.path.basename(file_path)} 다운로드",
                            data=f.read(),
                            file_name=os.path.basename(file_path),
                            mime='application/octet-stream',
                            key=f"download_button_{os.path.basename(file_path)}" # 고유한 key 추가
                        )

    with col1:
        st.subheader("📈 실시간 미세먼지 현황")
        
        if st.button("🔄 현재 데이터 조회"):
            if selected_sido:
                with st.spinner("데이터 조회 중..."):
                    air_quality_data = reporter.get_realtime_air_quality(selected_sido)
                    
                    if air_quality_data:
                        # 데이터프레임으로 변환
                        df = pd.DataFrame(air_quality_data)
                        
                        # 수치 데이터 변환
                        numeric_columns = ['pm10Value', 'pm25Value', 'khaiValue']
                        for col in numeric_columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                        
                        # 요약 통계
                        col1_1, col1_2, col1_3, col1_4 = st.columns(4)
                        
                        with col1_1:
                            avg_pm10 = df['pm10Value'].mean()
                            st.metric("평균 PM10", f"{avg_pm10:.1f} ㎍/㎥" if not pd.isna(avg_pm10) else "N/A")
                        
                        with col1_2:
                            avg_pm25 = df['pm25Value'].mean()
                            st.metric("평균 PM2.5", f"{avg_pm25:.1f} ㎍/㎥" if not pd.isna(avg_pm25) else "N/A")
                        
                        with col1_3:
                            max_pm10 = df['pm10Value'].max()
                            st.metric("최고 PM10", f"{max_pm10:.1f} ㎍/㎥" if not pd.isna(max_pm10) else "N/A")
                        
                        with col1_4:
                            measuring_stations = len(df)
                            st.metric("측정소 수", f"{measuring_stations}개")
                        
                        # 차트
                        if not df.empty:
                            # PM10 막대 차트
                            fig_pm10 = px.bar(
                                df.head(10), 
                                x='stationName', 
                                y='pm10Value',
                                title=f'{selected_sido} PM10 농도 (상위 10개 측정소)',
                                labels={'pm10Value': 'PM10 농도 (㎍/㎥)', 'stationName': '측정소명'}
                            )
                            fig_pm10.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(fig_pm10, use_container_width=True)
                            
                            # PM2.5 막대 차트
                            fig_pm25 = px.bar(
                                df.head(10), 
                                x='stationName', 
                                y='pm25Value',
                                title=f'{selected_sido} PM2.5 농도 (상위 10개 측정소)',
                                labels={'pm25Value': 'PM2.5 농도 (㎍/㎥)', 'stationName': '측정소명'},
                                color='pm25Value',
                                color_continuous_scale='Reds'
                            )
                            fig_pm25.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(fig_pm25, use_container_width=True)
                        
                        # 데이터 테이블
                        st.subheader("📋 상세 데이터")
                        display_columns = ['stationName', 'pm10Value', 'pm25Value', 
                                         'pm10Grade', 'pm25Grade', 'dataTime']
                        
                        display_df = df[display_columns].copy()
                        display_df.columns = ['측정소명', 'PM10', 'PM2.5', 'PM10등급', 'PM2.5등급', '측정시간']
                        
                        # 등급을 텍스트로 변환
                        for grade_col in ['PM10등급', 'PM2.5등급']:
                            display_df[grade_col] = display_df[grade_col].apply(
                                lambda x: reporter.get_air_quality_grade_text(x)
                            )
                        
                        st.dataframe(display_df, use_container_width=True)
                    else:
                        st.error("데이터를 불러올 수 없습니다.")

if __name__ == "__main__":
    main()