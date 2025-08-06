import requests
import pandas as pd
import json
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import traceback
import urllib.parse

class AirQualityReporter:
    import requests
import pandas as pd
import json
from datetime import datetime, timedelta
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import traceback
import urllib.parse

class AirQualityReporter:
    def __init__(self, api_key):
        """
        미세먼지 리포터 초기화
        
        Args:
            api_key (str): 에어코리아 Open API 키
        """
        self.api_key = api_key
        self.base_url = "http://apis.data.go.kr/B552584/ArpltnInforInqireSvc"
        
    def get_sido_list(self):
        """시도 목록 조회"""
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': self.api_key,
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': '서울',
            'ver': '1.0'
        }
        
        try:
            response = requests.get(url, params=params)
            if response.status_code == 200:
                # 주요 시도 목록 (실제 API에서는 동적으로 가져올 수 있음)
                sido_list = [
                    '서울', '부산', '대구', '인천', '광주', '대전', '울산', '세종',
                    '경기', '강원', '충북', '충남', '전북', '전남', '경북', '경남', '제주'
                ]
                return sido_list
        except Exception as e:
            print(f"시도 목록 조회 오류: {e}")
            return []
    
    def get_realtime_air_quality(self, sido_name):
        """실시간 대기질 데이터 조회"""
        # 먼저 XML로 시도
        xml_data = self._fetch_air_quality_xml(sido_name)
        if xml_data:
            return xml_data
        
        # XML 실패 시 JSON으로 시도
        return self._fetch_air_quality_json(sido_name)
    
    def _fetch_air_quality_xml(self, sido_name):
        """XML 형태로 대기질 데이터 조회"""
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': self.api_key,
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': sido_name,
            'ver': '1.0'
        }
        
        try:
            print(f"[XML] API 요청 URL: {url}")
            print(f"[XML] API 요청 파라미터: {params}")
            
            response = requests.get(url, params=params, timeout=30)
            print(f"[XML] 응답 상태 코드: {response.status_code}")
            print(f"[XML] Content-Type: {response.headers.get('content-type', 'Unknown')}")
            print(f"[XML] 응답 내용 (처음 300자): {response.text[:300]}")
            
            if response.status_code == 200:
                return self._parse_xml_response(response.text, sido_name)
            else:
                print(f"[XML] API 요청 실패. 상태 코드: {response.status_code}")
                return None
                
        except Exception as e:
            print(f"[XML] 요청 오류: {e}")
            return None
    
    def _fetch_air_quality_json(self, sido_name):
        """JSON 형태로 대기질 데이터 조회"""
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': self.api_key,
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': sido_name,
            'ver': '1.0'
        }
        
        try:
            print(f"[JSON] API 요청 URL: {url}")
            print(f"[JSON] API 요청 파라미터: {params}")
            
            response = requests.get(url, params=params, timeout=30)
            print(f"[JSON] 응답 상태 코드: {response.status_code}")
            print(f"[JSON] Content-Type: {response.headers.get('content-type', 'Unknown')}")
            print(f"[JSON] 응답 내용 (처음 300자): {response.text[:300]}")
            
            if response.status_code != 200:
                print(f"[JSON] API 요청 실패. 상태 코드: {response.status_code}")
                return None
            
            # 빈 응답 체크
            if not response.text.strip():
                print("[JSON] 빈 응답 수신")
                return None
            
            # JSON 파싱 시도
            try:
                data = response.json()
                print(f"[JSON] 파싱 성공: {type(data)}")
                
                if isinstance(data, dict) and 'response' in data:
                    response_data = data['response']
                    if 'body' in response_data and 'items' in response_data['body']:
                        items = response_data['body']['items']
                        if items:
                            print(f"[JSON] 아이템 수: {len(items)}")
                            return self._process_air_quality_data(items)
                        else:
                            print("[JSON] 아이템이 비어있음")
                    else:
                        print(f"[JSON] body/items가 없음: {response_data.keys()}")
                else:
                    print(f"[JSON] 예상과 다른 응답 구조: {type(data)}")
                
                return None
                
            except json.JSONDecodeError as json_err:
                print(f"[JSON] JSON 파싱 오류: {json_err}")
                print(f"[JSON] 응답 전체 내용: {response.text}")
                return None
                
        except requests.exceptions.Timeout:
            print("[JSON] 요청 타임아웃")
            return None
        except requests.exceptions.RequestException as req_err:
            print(f"[JSON] 네트워크 요청 오류: {req_err}")
            return None
        except Exception as e:
            print(f"[JSON] 알 수 없는 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _parse_xml_response(self, xml_text, sido_name):
        """XML 응답 파싱"""
        try:
            print(f"[XML Parser] XML 파싱 시작")
            print(f"[XML Parser] XML 내용 (처음 500자): {xml_text[:500]}")
            
            root = ET.fromstring(xml_text)
            print(f"[XML Parser] 루트 요소: {root.tag}")
            
            # 네임스페이스 처리
            namespaces = dict([node for _, node in ET.iterparse(BytesIO(xml_text.encode()), events=['start-ns'])])
            print(f"[XML Parser] 네임스페이스: {namespaces}")
            
            # 에러 코드 확인 (여러 가능한 경로)
            result_code_paths = ['.//resultCode', './/header/resultCode', './/{*}resultCode']
            result_code = None
            for path in result_code_paths:
                result_code = root.find(path)
                if result_code is not None:
                    break
            
            if result_code is not None:
                code = result_code.text
                print(f"[XML Parser] 결과 코드: {code}")
                
                if code != '00':
                    result_msg_paths = ['.//resultMsg', './/header/resultMsg', './/{*}resultMsg']
                    result_msg = None
                    for path in result_msg_paths:
                        result_msg = root.find(path)
                        if result_msg is not None:
                            break
                    
                    error_msg = result_msg.text if result_msg is not None else "알 수 없는 오류"
                    print(f"[XML Parser] API 에러: {code} - {error_msg}")
                    return None
            
            # items 데이터 추출 (여러 가능한 경로)
            items_paths = ['.//items/item', './/item', './/{*}items/{*}item', './/{*}item']
            items = []
            
            for path in items_paths:
                found_items = root.findall(path)
                if found_items:
                    print(f"[XML Parser] {len(found_items)}개 아이템을 경로 '{path}'에서 발견")
                    for item in found_items:
                        item_data = {}
                        for child in item:
                            # 네임스페이스 제거한 태그명 사용
                            tag_name = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                            item_data[tag_name] = child.text
                        items.append(item_data)
                    break
            
            print(f"[XML Parser] 총 {len(items)}개 아이템 추출")
            
            if items:
                # 첫 번째 아이템 샘플 출력
                if len(items) > 0:
                    print(f"[XML Parser] 첫 번째 아이템 키: {list(items[0].keys())}")
                return self._process_air_quality_data(items)
            else:
                print(f"[XML Parser] '{sido_name}' 지역의 데이터를 찾을 수 없습니다.")
                # XML 구조 디버깅
                print("[XML Parser] XML 구조 분석:")
                self._debug_xml_structure(root, 0, 3)  # 최대 3레벨까지 출력
                return None
                
        except ET.ParseError as parse_err:
            print(f"[XML Parser] XML 파싱 오류: {parse_err}")
            print(f"[XML Parser] 오류 위치 근처 내용: {xml_text[max(0, parse_err.position[1]-100):parse_err.position[1]+100]}")
            return None
        except Exception as e:
            print(f"[XML Parser] XML 응답 처리 오류: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _debug_xml_structure(self, element, level, max_level):
        """XML 구조 디버깅용 메소드"""
        if level > max_level:
            return
        
        indent = "  " * level
        tag_name = element.tag.split('}')[-1] if '}' in element.tag else element.tag
        
        if element.text and element.text.strip():
            print(f"{indent}{tag_name}: {element.text.strip()}")
        else:
            print(f"{indent}{tag_name}")
        
        for child in list(element)[:3]:  # 각 레벨에서 최대 3개 자식만 출력
            self._debug_xml_structure(child, level + 1, max_level)

    def test_api_connection(self):
        """API 연결 테스트"""
        print("=" * 50)
        print("API 연결 테스트 시작")
        print("=" * 50)
        
        # 1. 기본 연결 테스트
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        test_params = {
            'serviceKey': self.api_key,
            'numOfRows': '1',
            'pageNo': '1',
            'sidoName': '서울',
            'ver': '1.0'
        }
        
        try:
            print(f"테스트 URL: {url}")
            print(f"테스트 파라미터: {test_params}")
            
            response = requests.get(url, params=test_params, timeout=10)
            print(f"응답 상태 코드: {response.status_code}")
            print(f"응답 헤더: {dict(response.headers)}")
            print(f"응답 내용 (처음 500자):\n{response.text[:500]}")
            
            if response.status_code == 200:
                # API 키 유효성 확인
                if 'INVALID' in response.text.upper() or 'ERROR' in response.text.upper():
                    print("⚠️ API 키가 유효하지 않을 가능성이 있습니다.")
                    return False
                
                # XML 파싱 테스트
                if response.text.strip().startswith('<?xml') or '<' in response.text:
                    try:
                        root = ET.fromstring(response.text)
                        result_code = root.find('.//resultCode')
                        if result_code is not None:
                            if result_code.text == '00':
                                print("✅ API 연결 및 인증 성공!")
                                return True
                            else:
                                result_msg = root.find('.//resultMsg')
                                error_msg = result_msg.text if result_msg is not None else "알 수 없는 오류"
                                print(f"❌ API 오류: {result_code.text} - {error_msg}")
                                return False
                        else:
                            print("⚠️ 결과 코드를 찾을 수 없습니다.")
                            return False
                    except ET.ParseError as e:
                        print(f"❌ XML 파싱 오류: {e}")
                        return False
                else:
                    print("⚠️ 예상과 다른 응답 형식입니다.")
                    return False
            else:
                print(f"❌ HTTP 오류: {response.status_code}")
                return False
                
        except requests.exceptions.Timeout:
            print("❌ 요청 타임아웃")
            return False
        except requests.exceptions.RequestException as e:
            print(f"❌ 네트워크 오류: {e}")
            return False
        except Exception as e:
            print(f"❌ 알 수 없는 오류: {e}")
            import traceback
            traceback.print_exc()
            return False
        """대기질 예보 정보 조회"""
        url = f"{self.base_url}/getMinuDustFrcstDspth"
        params = {
            'serviceKey': self.api_key,
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'searchDate': search_date,
            'InformCode': 'PM10'
        }
        
        try:
            response = requests.get(url, params=params)
            if response.status_code == 200:
                data = response.json()
                if 'response' in data and 'body' in data['response']:
                    items = data['response']['body']['items']
                    return items
            return None
        except Exception as e:
            print(f"대기질 예보 조회 오류: {e}")
            return None
    
    def _process_air_quality_data(self, items):
        """대기질 데이터 처리"""
        processed_data = []
        for item in items:
            processed_item = {
                'stationName': item.get('stationName', ''),
                'sidoName': item.get('sidoName', ''),
                'pm10Value': item.get('pm10Value', ''),
                'pm25Value': item.get('pm25Value', ''),
                'pm10Grade': item.get('pm10Grade', ''),
                'pm25Grade': item.get('pm25Grade', ''),
                'so2Value': item.get('so2Value', ''),
                'coValue': item.get('coValue', ''),
                'o3Value': item.get('o3Value', ''),
                'no2Value': item.get('no2Value', ''),
                'dataTime': item.get('dataTime', ''),
                'khaiValue': item.get('khaiValue', ''),
                'khaiGrade': item.get('khaiGrade', '')
            }
            processed_data.append(processed_item)
        return processed_data


    def __init__(self, api_key):
        """
        미세먼지 리포터 초기화
        
        Args:
            api_key (str): 에어코리아 Open API 키
        """
        self.api_key = api_key
        self.base_url = "http://apis.data.go.kr/B552584/ArpltnInforInqireSvc"
        
    def get_sido_list(self):
        """
        시도 목록을 조회합니다.
        
        Returns:
            list: 시도명 (예: '서울', '부산') 리스트. 조회 실패 시 빈 리스트 반환.
        """
        url = f"{self.base_url}/getCtprvnRltmMesureDnsty"
        params = {
            'serviceKey': urllib.parse.unquote(self.api_key),
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1'
        }
        
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status() # HTTP 오류 발생 시 예외 발생
            data = response.json()
            
            if data and 'response' in data and 'body' in data['response'] and 'items' in data['response']['body']:
                sido_names = sorted(list(set([item['sidoName'] for item in data['response']['body']['items'] if 'sidoName' in item])))
                return sido_names
            else:
                st.error("시도 목록 조회 결과에 유효한 데이터가 없습니다.")
                return []
        except requests.exceptions.Timeout:
            st.error("API 요청 시간 초과: 시도 목록을 불러올 수 없습니다.")
            return []
        except requests.exceptions.ConnectionError:
            st.error("API 연결 오류: 네트워크 상태를 확인하거나 잠시 후 다시 시도해주세요.")
            return []
        except requests.exceptions.RequestException as e:
            st.error(f"시도 목록 조회 중 오류 발생: {e}")
            return []
        except json.JSONDecodeError:
            st.error("시도 목록 API 응답이 유효한 JSON 형식이 아닙니다.")
            return []

    def get_realtime_air_quality(self, sido_name):
        """
        특정 시도의 실시간 대기질 정보를 조회합니다.
        
        Args:
            sido_name (str): 조회할 시도명 (예: '서울').
            
        Returns:
            list: 각 측정소의 대기질 데이터를 담은 딕셔너리 리스트. 조회 실패 시 None 반환.
        """
        url = f"{self.base_url}/getCtprvnMesureLIst"
        params = {
            'serviceKey': urllib.parse.unquote(self.api_key),
            'returnType': 'json',
            'numOfRows': '100',
            'pageNo': '1',
            'sidoName': sido_name,
            'ver': '1.3' # 최신 버전
        }
        
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            if data and 'response' in data and 'body' in data['response'] and 'items' in data['response']['body']:
                return data['response']['body']['items']
            else:
                st.warning(f"{sido_name}의 실시간 대기질 데이터를 찾을 수 없습니다.")
                return None
        except requests.exceptions.Timeout:
            st.error(f"API 요청 시간 초과: {sido_name}의 대기질 데이터를 불러올 수 없습니다.")
            return None
        except requests.exceptions.ConnectionError:
            st.error("API 연결 오류: 네트워크 상태를 확인하거나 잠시 후 다시 시도해주세요.")
            return None
        except requests.exceptions.RequestException as e:
            st.error(f"대기질 데이터 조회 중 오류 발생: {e}")
            return None
        except json.JSONDecodeError:
            st.error("대기질 데이터 API 응답이 유효한 JSON 형식이 아닙니다.")
            return None

    def test_api_connection(self):
        """
        API 연결 및 API 키 유효성을 테스트합니다.
        간단한 시도 목록 조회를 통해 확인합니다.
        
        Returns:
            bool: 연결 및 키가 유효하면 True, 그렇지 않으면 False.
        """
        try:
            # 시도 목록을 조회하는 것으로 테스트
            sido_list = self.get_sido_list()
            if sido_list: # 시도 목록이 비어있지 않으면 성공으로 간주
                return True
            else:
                st.warning("API 키가 유효하지 않거나 데이터가 없습니다.")
                return False
        except Exception as e:
            st.error(f"API 연결 테스트 중 예외 발생: {e}")
            return False

    def create_pdf_report(self, data, sido_name, filename):
        """
        대기질 데이터를 PDF 리포트로 생성합니다.
        
        Args:
            data (list): 대기질 데이터 리스트.
            sido_name (str): 시도명.
            filename (str): 저장할 PDF 파일명.
        """
        # 한글 폰트 등록
        try:
            pdfmetrics.registerFont(TTFont('MalgunGothic', 'MalgunGothic.ttf'))
        except Exception:
            st.error("🚨 'MalgunGothic.ttf' 폰트를 찾을 수 없습니다. 리포트가 깨질 수 있습니다. 폰트 파일을 스크립트와 같은 경로에 넣어주세요.")
            pdfmetrics.registerFont(TTFont('MalgunGothic', 'malgun.ttf')) # 대체 폰트명 시도

        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # 기본 스타일 설정 (한글 폰트 적용)
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        styles.add(ParagraphStyle(name='TitleStyle', fontSize=24, leading=28, alignment=TA_CENTER, fontName='MalgunGothic'))
        styles.add(ParagraphStyle(name='Heading1', fontSize=18, leading=22, alignment=TA_LEFT, fontName='MalgunGothic'))
        styles.add(ParagraphStyle(name='Normal', fontSize=10, leading=12, alignment=TA_LEFT, fontName='MalgunGothic'))
        styles.add(ParagraphStyle(name='Footer', fontSize=8, leading=10, alignment=TA_RIGHT, fontName='MalgunGothic'))

        story = []

        # 제목
        story.append(Paragraph(f"{sido_name} 실시간 미세먼지 리포트", styles['TitleStyle']))
        story.append(Spacer(1, 0.2 * inch))

        # 리포트 정보
        story.append(Paragraph(f"생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H시 %M분')}", styles['Normal']))
        story.append(Paragraph(f"대상지역: {sido_name}", styles['Normal']))
        story.append(Paragraph(f"측정소 수: {len(data)}개", styles['Normal']))
        story.append(Spacer(1, 0.2 * inch))

        # 데이터 테이블
        table_data = [['측정소명', 'PM10 (㎍/㎥)', 'PM2.5 (㎍/㎥)', 'PM10 등급', 'PM2.5 등급', '측정시간']]
        for item in data:
            row = [
                item.get('stationName', 'N/A'),
                item.get('pm10Value', 'N/A'),
                item.get('pm25Value', 'N/A'),
                self.get_air_quality_grade_text(item.get('pm10Grade', 'N/A')),
                self.get_air_quality_grade_text(item.get('pm25Grade', 'N/A')),
                item.get('dataTime', 'N/A')
            ]
            table_data.append(row)

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')), # 헤더 배경색 (녹색 계열)
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), # 헤더 텍스트색
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'MalgunGothic'), # 헤더 폰트
            ('FONTNAME', (0, 1), (-1, -1), 'MalgunGothic'), # 바디 폰트
            ('FONTSIZE', (0,0), (-1,-1), 9), # 폰트 크기
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige), # 바디 배경색
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.5 * inch))

        # 푸터
        story.append(Paragraph("데이터 출처: 에어코리아 Open API", styles['Footer']))

        try:
            doc.build(story)
            st.success(f"✅ PDF 리포트 '{filename}'이 성공적으로 생성되었습니다.")
        except Exception as e:
            st.error(f"❌ PDF 리포트 생성 중 오류 발생: {e}")

    def create_excel_report(self, data, sido_name, filename):
        """
        대기질 데이터를 Excel 리포트로 생성합니다.
        
        Args:
            data (list): 대기질 데이터 리스트.
            sido_name (str): 시도명.
            filename (str): 저장할 Excel 파일명.
        """
        df = pd.DataFrame(data)
        
        # 필요한 컬럼만 선택하고 순서 변경
        columns_to_keep = ['stationName', 'pm10Value', 'pm25Value', 'pm10Grade', 'pm25Grade', 'dataTime']
        df = df[columns_to_keep]

        # 컬럼명 한글로 변경
        df.columns = ['측정소명', 'PM10 (㎍/㎥)', 'PM2.5 (㎍/㎥)', 'PM10 등급', 'PM2.5 등급', '측정시간']
        
        # 등급을 텍스트로 변환
        for grade_col in ['PM10 등급', 'PM2.5 등급']:
            df[grade_col] = df[grade_col].apply(lambda x: self.get_air_quality_grade_text(x))

        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='대기질 현황', index=False)
                
                workbook = writer.book
                sheet = writer.sheets['대기질 현황']
                
                # 헤더 스타일 적용
                header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    sheet.column_dimensions[chr(64 + col_idx)].width = 15 # 컬럼 너비 자동 조정
                
                # 데이터 부분 폰트 설정
                data_font = Font(name='맑은 고딕')
                for row_idx in range(2, len(df) + 2):
                    for col_idx in range(1, len(df.columns) + 1):
                        sheet.cell(row=row_idx, column=col_idx).font = data_font

            st.success(f"✅ Excel 리포트 '{filename}'이 성공적으로 생성되었습니다.")
        except Exception as e:
            st.error(f"❌ Excel 리포트 생성 중 오류 발생: {e}")

    def get_air_quality_grade_text(self, grade):
        """
        대기질 등급 코드에 따른 텍스트를 반환합니다.
        
        Args:
            grade (str or int): 대기질 등급 코드.
            
        Returns:
            str: 등급 텍스트.
        """
        grade_map = {
            '1': '좋음', '2': '보통', '3': '나쁨', '4': '매우 나쁨',
            1: '좋음', 2: '보통', 3: '나쁨', 4: '매우 나쁨'
        }
        return grade_map.get(grade, '알 수 없음')

    def send_email_report(self, sender_email, sender_password, recipient_emails, subject, body, attachments):
        """
        이메일로 리포트를 발송합니다.
        
        Args:
            sender_email (str): 발송자 이메일 주소.
            sender_password (str): 발송자 이메일 비밀번호 (앱 비밀번호 사용 권장).
            recipient_emails (list): 수신자 이메일 주소 리스트.
            subject (str): 이메일 제목.
            body (str): 이메일 본문.
            attachments (list): 첨부할 파일 경로 리스트.
            
        Returns:
            bool: 이메일 발송 성공 여부.
        """
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ", ".join(recipient_emails)
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'plain'))
        
        for file_path in attachments:
            try:
                part = MIMEBase('application', 'octet-stream')
                with open(file_path, 'rb') as file:
                    part.set_payload(file.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(file_path)}")
                msg.attach(part)
            except Exception as e:
                st.error(f"첨부 파일 {file_path}를 추가하는 데 실패했습니다: {e}")
                return False
                
        try:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(sender_email, sender_password)
                smtp.send_message(msg)
            return True
        except Exception as e:
            st.error(f"이메일 발송 중 오류 발생: {e}")
            traceback.print_exc() # 상세 오류 출력
            return False

# Streamlit 앱
API_KEY = "IlKf+/s38dEMDB7W9oToxSVtk9Cre94VFL3jEPuj/Gf4LU7jE9CfQBxrkymu3+2Ix2JfgnCSOuitIzZgfGIjWw==" # 여기에 실제 API 키를 입력하세요!

# reportlab 관련 스타일 임포트 추가 (PDF 생성을 위해 필요)
try:
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
except ImportError:
    st.error("ReportLab 스타일을 임포트할 수 없습니다. ReportLab이 설치되었는지 확인해주세요.")
    ParagraphStyle = type('ParagraphStyle', (object,), {}) # 더미 클래스
    TA_CENTER, TA_LEFT, TA_RIGHT = None, None, None

def main():
    st.set_page_config(layout="wide", page_title="미세먼지 리포트 생성기")

    st.title("🌱 실시간 미세먼지 리포트 생성기")
    st.markdown("---")

    if 'api_connection_ok' not in st.session_state:
        st.session_state.api_connection_ok = None

    if 'air_quality_data' not in st.session_state:
        st.session_state.air_quality_data = None
    if 'selected_sido_for_data' not in st.session_state:
        st.session_state.selected_sido_for_data = None
    if 'report_files' not in st.session_state:
        st.session_state.report_files = []
    if 'report_generated_success' not in st.session_state:
        st.session_state.report_generated_success = False
    if 'show_completion_message_done' not in st.session_state:
        st.session_state.show_completion_message_done = False

    # 사이드바에서 설정
    with st.sidebar:
        st.header("설정")

        reporter = AirQualityReporter(API_KEY)

        if st.session_state.api_connection_ok is None:
            status_placeholder = st.empty()
            with status_placeholder:
                with st.spinner("API 연결 및 키 유효성 확인 중..."):
                    test_result = reporter.test_api_connection()
                    st.session_state.api_connection_ok = test_result
            status_placeholder.empty()

        if not st.session_state.api_connection_ok:
            st.error("🚨 API 연결에 실패했거나 API 키가 유효하지 않습니다.")
            st.info("💡 API 키(`API_KEY` 변수)를 다시 확인하거나, 네트워크 연결 상태를 점검해주세요. 이 문제가 해결되지 않으면 데이터 조회가 불가능합니다.")

        sido_list = reporter.get_sido_list()
        selected_sido = st.selectbox("지역 선택", sido_list, key="sido_selector")

        report_types = st.multiselect(
            "생성할 리포트 타입",
            ["PDF", "Excel"],
            default=["PDF"]
        )

        st.subheader("이메일 발송 설정")
        send_email = st.checkbox("이메일로 리포트 발송")

        if send_email:
            sender_email = st.text_input("발송자 이메일")
            sender_password = st.text_input("이메일 비밀번호", type="password")
            recipient_emails = st.text_input(
                "수신자 이메일 (콤마로 구분)",
                placeholder="email1@example.com, email2@example.com"
            )

    # 메인 콘텐츠
    col1, col2 = st.columns([2, 1])

    with col2:
        if st.button("📊 데이터 조회 및 리포트 생성", type="primary"):
            st.session_state.report_generated_success = False
            st.session_state.show_completion_message_done = False # 완료 메시지 초기화

            should_show_progress = True # 이 줄을 추가합니다.

            if not selected_sido:
                st.error("지역을 선택해주세요")
                should_show_progress = False # 이 줄로 변경 (기존 return 제거)

            elif not report_types: # 'if'를 'elif'로 변경합니다.
                st.error("⚠️ 리포트 파일 형식을 선택해주세요 (PDF 또는 Excel).")
                should_show_progress = False # 이 줄로 변경 (기존 return 제거)

            elif not st.session_state.api_connection_ok: # 'if'를 'elif'로 변경합니다.
                st.error("🚨 API 연결 문제가 해결되지 않아 데이터를 조회할 수 없습니다.")
                should_show_progress = False

            if should_show_progress: # 이 if 블록을 추가하고, 아래의 모든 코드를 이 안에 들여씁니다.
                # 프로그레스 바는 실제 작업이 시작될 때만 초기화
                progress_bar = st.progress(0)
                status_text = st.empty()

                # 데이터 조회
                status_text.text("실시간 대기질 데이터 조회 중...")
                progress_bar.progress(25)

                current_air_quality_data = reporter.get_realtime_air_quality(selected_sido)

                if not current_air_quality_data:
                    st.error("❌ 데이터를 불러올 수 없습니다.")
                    st.info("🔍 문제 해결 방법:\n1. API 키가 올바른지 확인\n2. 네트워크 연결 상태 확인\n3. API 키 승인 상태 확인\n4. 다른 지역 선택해보기")
                    st.session_state.report_generated_success = False
                    st.session_state.show_completion_message_done = False
                    progress_bar.empty() # 이 줄 추가
                    status_text.empty() # 이 줄 추가
                else:
                    # session_state에 데이터 저장 (성공했을 때만)
                    st.session_state.air_quality_data = current_air_quality_data
                    st.session_state.selected_sido_for_data = selected_sido

                    # 리포트 생성
                    status_text.text("리포트 생성 중...")
                    progress_bar.progress(50)

                    new_report_files = []

                    if "PDF" in report_types:
                        pdf_path = f"{selected_sido}_미세먼지_리포트_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
                        reporter.create_pdf_report(current_air_quality_data, selected_sido, pdf_path)
                        new_report_files.append(pdf_path)

                    if "Excel" in report_types:
                        excel_path = f"{selected_sido}_미세먼지_리포트_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                        reporter.create_excel_report(current_air_quality_data, selected_sido, excel_path)
                        new_report_files.append(excel_path)

                    st.session_state.report_files = new_report_files

                    progress_bar.progress(75)

                    # 이메일 발송
                    if send_email and sender_email and sender_password and recipient_emails:
                        status_text.text("이메일 발송 중...")
                        recipient_list = [email.strip() for email in recipient_emails.split(',')]

                        subject = f"[{selected_sido}] 미세먼지 현황 리포트 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                        body = f"""
안녕하세요.

{selected_sido} 지역의 미세먼지 현황 리포트를 첨부파일로 보내드립니다.

생성일시: {datetime.now().strftime('%Y년 %m월 %d일 %H시 %M분')}
대상지역: {selected_sido}
측정소 수: {len(current_air_quality_data)}개

감사합니다.
                        """

                        email_result = reporter.send_email_report(
                            sender_email, sender_password, recipient_list,
                            subject, body, st.session_state.report_files
                        )

                        if email_result:
                            st.success("✅ 이메일이 성공적으로 발송되었습니다!")
                        else:
                            st.error("❌ 이메일 발송에 실패했습니다.")

                    progress_bar.progress(100)
                    status_text.text("완료!")

                    if st.session_state.report_files:
                        st.session_state.report_generated_success = True
                        st.session_state.show_completion_message_done = True

        if st.session_state.show_completion_message_done:
            st.info("완료!")

        if st.session_state.report_generated_success and st.session_state.report_files:
            st.success(f"✅ 리포트가 성공적으로 생성되었습니다!")

        # 다운로드 버튼 (session_state에 파일이 있을 경우 항상 표시)
        if st.session_state.report_files:
            st.subheader("⬇️ 리포트 다운로드")
            for file_path in st.session_state.report_files:
                if os.path.exists(file_path):
                    with open(file_path, 'rb') as f:
                        st.download_button(
                            label=f"📁 {os.path.basename(file_path)} 다운로드",
                            data=f.read(),
                            file_name=os.path.basename(file_path),
                            mime='application/octet-stream',
                            key=f"download_button_{os.path.basename(file_path)}"
                        )

    with col1:
        st.subheader("📈 실시간 미세먼지 현황")

        if st.button("🔄 현재 데이터 조회", key="refresh_data_button"):
            if not st.session_state.api_connection_ok:
                st.error("🚨 API 연결 문제가 해결되지 않아 데이터를 조회할 수 없습니다.")
                return

            if selected_sido:
                with st.spinner("데이터 조회 중..."):
                    current_air_quality_data = reporter.get_realtime_air_quality(selected_sido)

                    if not current_air_quality_data:
                        st.error("❌ 데이터를 불러올 수 없습니다.")
                    else:
                        st.session_state.air_quality_data = current_air_quality_data
                        st.session_state.selected_sido_for_data = selected_sido


        # air_quality_data가 존재하면 항상 차트와 테이블을 표시
        if st.session_state.air_quality_data:
            df = pd.DataFrame(st.session_state.air_quality_data)

            if st.session_state.selected_sido_for_data != selected_sido:
                st.info(f"현재 표시되는 데이터는 '{st.session_state.selected_sido_for_data}' 지역의 데이터입니다. '{selected_sido}'에 대한 최신 데이터를 조회하려면 '현재 데이터 조회' 버튼을 눌러주세요.")

            numeric_columns = ['pm10Value', 'pm25Value', 'khaiValue']
            for col in numeric_columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

            col1_1, col1_2, col1_3, col1_4 = st.columns(4)

            with col1_1:
                avg_pm10 = df['pm10Value'].mean()
                st.metric("평균 PM10", f"{avg_pm10:.1f} ㎍/㎥" if not pd.isna(avg_pm10) else "N/A")

            with col1_2:
                avg_pm25 = df['pm25Value'].mean()
                st.metric("평균 PM2.5", f"{avg_pm25:.1f} ㎍/㎥" if not pd.isna(avg_pm25) else "N/A")

            with col1_3:
                max_pm10 = df['pm10Value'].max()
                st.metric("최고 PM10", f"{max_pm10:.1f} ㎍/㎥" if not pd.isna(max_pm10) else "N/A")

            with col1_4:
                measuring_stations = len(df)
                st.metric("측정소 수", f"{measuring_stations}개")

            if not df.empty:
                fig_pm10 = px.bar(
                    df.head(10),
                    x='stationName',
                    y='pm10Value',
                    title=f'{st.session_state.selected_sido_for_data} PM10 농도 (상위 10개 측정소)',
                    labels={'pm10Value': 'PM10 농도 (㎍/㎥)', 'stationName': '측정소명'}
                )
                fig_pm10.update_layout(xaxis_tickangle=-45)
                st.plotly_chart(fig_pm10, use_container_width=True)

                fig_pm25 = px.bar(
                    df.head(10),
                    x='stationName',
                    y='pm25Value',
                    title=f'{st.session_state.selected_sido_for_data} PM2.5 농도 (상위 10개 측정소)',
                    labels={'pm25Value': 'PM2.5 농도 (㎍/㎥)', 'stationName': '측정소명'},
                    color='pm25Value',
                    color_continuous_scale='Reds'
                )
                fig_pm25.update_layout(xaxis_tickangle=-45)
                st.plotly_chart(fig_pm25, use_container_width=True)

            st.subheader("📋 상세 데이터")
            display_columns = ['stationName', 'pm10Value', 'pm25Value',
                                 'pm10Grade', 'pm25Grade', 'dataTime']

            display_df = df[display_columns].copy()
            display_df.columns = ['측정소명', 'PM10', 'PM2.5', 'PM10등급', 'PM2.5등급', '측정시간']

            for grade_col in ['PM10등급', 'PM2.5등급']:
                display_df[grade_col] = display_df[grade_col].apply(
                    lambda x: reporter.get_air_quality_grade_text(x)
                )

            st.dataframe(display_df, use_container_width=True)
        else:
            st.info("좌측의 '데이터 조회 및 리포트 생성' 또는 상단의 '현재 데이터 조회' 버튼을 눌러 데이터를 불러오세요.")

if __name__ == "__main__":
    main()