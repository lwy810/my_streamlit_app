import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font

def create_simple_formatted_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TestSheet"

    # A1 셀에 텍스트 쓰기
    sheet['A1'] = "테스트 텍스트"

    # A1 셀에 서식 적용
    cell = sheet['A1']
    
    # 테두리
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    cell.border = thin_border
    
    # 배경색 (노란색)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell.fill = yellow_fill
    
    # 중앙 정렬
    center_aligned_text = Alignment(horizontal="center", vertical="center")
    cell.alignment = center_aligned_text
    
    # 굵은 글씨
    bold_font = Font(bold=True)
    cell.font = bold_font

    # 파일 저장
    file_name = "simple_formatted_test.xlsx"
    workbook.save(file_name)
    print(f"'{file_name}' 파일이 생성되었습니다. 서식이 적용되었는지 확인해주세요.")

if __name__ == "__main__":
    create_simple_formatted_excel()